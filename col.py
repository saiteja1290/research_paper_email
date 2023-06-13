from openpyxl import load_workbook
def get_number_of_rows(excel_file, sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Count the number of non-empty rows
    number_of_rows = sum(1 for _ in ws.rows)

    return number_of_rows
def get_row_data(excel_file, sheet_name, row):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    row_data = []
    for cell in ws[row]:
        row_data.append(cell.value)

    return row_data

# Example usage
excel_file = 'registrations.xlsx'
sheet_name = 'Sheet'
row = 2
for i in range(get_number_of_rows(excel_file, sheet_name)-1):

    row_data = get_row_data(excel_file, sheet_name, row)
    row += 1
    print(row_data)
