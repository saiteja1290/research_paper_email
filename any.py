from bs4 import BeautifulSoup
import requests
import re
import random
from openpyxl import load_workbook
import os
import schedule
import time
import smtplib
from email.message import EmailMessage
from email.utils import formataddr
from pathlib import Path
from dotenv import load_dotenv

# -----------------------------------------------------
PORT = 587

EMAIL_SERVER = "smtp-mail.outlook.com"

current_dir = Path(__file__).resolve(
).parent if "__file__" in locals() else Path.cwd()
envars = current_dir / ".env"
load_dotenv(envars)

# Read environment variables
sender_email = os.getenv("EMAIL")
password_email = os.getenv("PASSWORD")
# ------------------------------------------------------
def send_email(subject, receiver_email, name, datas):
    # Create the base text message.
    msg = EmailMessage()
    msg["Subject"] = subject
    msg["From"] = formataddr(("Coding Is Fun Corp.", f"{sender_email}"))
    msg["To"] = receiver_email
    msg["BCC"] = sender_email
    messageraedi = f"""\
        Hi {name},
        Here is your weekly Research paper recommendations based on your interests -
        {data}
        Best regards,
        Organization Tenet
        """
    msg.set_content(
        messageraedi
    )
    # Add the html version.  This converts the message into a multipart/alternative
    # container, with the original text message as the first part and the new html
    # message as the second part.
    msg.add_alternative(
        f"""\
    <html>
      <body>
        <p>{messageraedi}</p>
      </body>
    </html>
    """,
        subtype="html",
    )

    with smtplib.SMTP(EMAIL_SERVER, PORT) as server:
        server.starttls()
        server.login(sender_email, password_email)
        server.sendmail(sender_email, receiver_email, msg.as_string())


def read_text_file(file_path):
    with open(file_path, 'r') as file:
        text = file.read()
    return text
def clean_sentence(sentence):
    # Remove all non-alphabetic, non-numeric, and non-space characters
    cleaned_sentence = re.sub(r'[^a-zA-Z0-9 ]', '', sentence)

    return cleaned_sentence
def link_to_pdfs_and_titles(domain, subdomain, name, page = 1):
    domain = domain.replace(" ", "+")
    subdomain = subdomain.replace(" ", "+")
    titles = []
    titles2 = []
    pdf_links = []
    url = f"https://link.springer.com/search/page/{page}?facet-discipline=%22{domain}%22&facet-sub-discipline=%22{subdomain}%22&showAll=false"

    response = requests.get(url)
    html_content = response.content
    soup = BeautifulSoup(html_content, "html.parser")


    target_links = soup.find_all("a", class_="webtrekk-track pdf-link")
    # target_links_title = soup.find_all("a", class_="title")
    for link in target_links:
        if(link.get("doi")):    
            # titles = link.text
            href = link.get("doi")
            title = link.get("aria-label")
            pdf_links.append("https://link.springer.com//content/pdf/"+href)
            titles.append(title)
    # print(pdf_links)
    # print(title)
    for title in titles:
        modified_text = re.sub(r"Download PDF \(\d+ KB\) - ", "", title)
        # title = modified_text
        titles2.append(modified_text)
    # print(titles2)

    file_path = f"text_files/{name}.txt"  # Replace with the desired file path

    # Open the file in write mode
    with open(file_path, "w") as file:
        for i in range(min(len(titles2), len(pdf_links))):
            file.write(clean_sentence(titles2[i].replace('\xa0', '')) + ": "+ pdf_links[i]+"\n\n")



def get_number_of_rows(excel_file, sheet_name):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    # Count the number of non-empty rows
    number_of_rows = sum(1 for _ in ws.rows)

    return number_of_rows
def get_row_data(excel_file, sheet_name, row):
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]

    row_data = []
    for cell in ws[row]:
        row_data.append(cell.value)

    return row_data

# Example usage
excel_file = 'registrations.xlsx'
sheet_name = 'Sheet'
row = 2
for i in range(get_number_of_rows(excel_file, sheet_name)-1):

    row_data = get_row_data(excel_file, sheet_name, row)
    domain = row_data[1]
    subdomain = row_data[3]
    name = row_data[0]
    email = row_data[2]
    row += 1
    link_to_pdfs_and_titles(domain, subdomain, name)
    data = read_text_file(f'text_files/{name}.txt')
    send_email(
        subject="Your Weekly Research Paper recommendation",
        name=f"{name}",
        receiver_email=f"{email}",
        datas= data,
    )