from flask import Flask, render_template, request, url_for
from openpyxl import Workbook
from openpyxl import load_workbook
from bs4 import BeautifulSoup
import requests
import re
import random

app = Flask(__name__, static_folder='static')

# This function saves the data in excel sheet
def save_to_excel(name, interests, email, subdomain):
    excel_file = 'registrations.xlsx'
    
    try:
        # Load the existing workbook
        wb = load_workbook(excel_file)
        ws = wb.active
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        ws['A1'] = 'Name'
        ws['B1'] = 'Interests'
        ws['C1'] = 'Email'
        ws['D1'] = 'Subdomain'
    
    ws.append([name, interests, email, subdomain])
    
    wb.save(excel_file)


@app.route("/")
def index():
    return render_template('index.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        name = request.form['name']
        interests = request.form['interests']
        email = request.form['email']
        subdomain = request.form['subdomain']

        if not name or not interests or not email or not subdomain:
            return '<h2 style="text-align:center">Please fill in all fields.</h2>'

        save_to_excel(name, interests, email, subdomain)

        return '<h2 style="text-align:center">Registration successful!</h2>'

    else:
        return render_template('register.html')


if __name__ == '__main__':
    app.run(debug=True)

